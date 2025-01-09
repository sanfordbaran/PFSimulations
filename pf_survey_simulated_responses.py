#!/usr/bin/env python
# coding: utf-8

import os
from openai import OpenAI
import pandas as pd
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import argparse as ap

# Load environment variables

# Constants
MODEL = "gpt-4o-mini"
MAX_TOKENS = 2000
TEMP0 = 0.0
TEMP14 = 1.4
TEMP2 = 2.0

# Define global variables
global num_plus_100, num_minus_100 
num_plus_100 = 0
num_minus_100 = 0

# Read tokens/keys from .env file
from dotenv import load_dotenv, find_dotenv
_ = load_dotenv('.env') # read local .env file

# Initialize the OpenAI client as a constant (CLIENT)
CLIENT = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))


def setup_logging(experiment_num):
    logging.basicConfig(
    filename=f'E{experiment_num}_pf_survey_simulated_responses.log',  # Specify the log file name
    level=logging.INFO,         # Set the logging level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    logging.getLogger("httpx").setLevel(logging.WARNING)



def get_completion_and_token_count(messages, 
                                   model=MODEL,
                                   temperature=TEMP0, 
                                   max_tokens=2000):
    """
    Generate a chat completion response and provide token count information.

    This function uses the OpenAI API to generate a chat completion response based on a list
    of conversation messages. It sends the messages to the specified model and receives a response
    with the degree of randomness controlled by the 'temperature' parameter and a maximum token limit
    set by 'max_tokens'. Additionally, it provides information about the token count used for the completion.

    Parameters:
    messages (list): A list of message objects, typically containing user and system messages.
    model (str): The name of the OpenAI model to use (default is 'gpt-3.5-turbo').
    temperature (float): The degree of randomness of the model's output (default is 0, less random).
    max_tokens (int): The maximum number of tokens the model can output (default is 2000).

    Returns:
    tuple: A tuple containing two elements:
        - str: The generated completion content as a string.
        - dict: A dictionary containing token count information:
            - 'prompt_tokens': The number of tokens used for the prompt.
            - 'completion_tokens': The number of tokens used for the completion.
            - 'total_tokens': The total number of tokens used.
    """

    #logging.info("Start ChatCompletion...")
    response = CLIENT.chat.completions.create(
        model=model,
        messages=messages,
        temperature=temperature, 
        max_tokens=max_tokens,
    )
    #logging.info("ChatCompletion returned")

    content = response.choices[0].message.content

    token_dict = {'prompt_tokens':response.usage.prompt_tokens,
                    'completion_tokens':response.usage.completion_tokens,
                    'total_tokens':response.usage.total_tokens
    }

    return content, token_dict



def get_score_and_rationale_using_gpt(statement, group_type, group_type_description, temp=TEMP0):
    """
    Simulate how a particular group type would evaluate a single statement from one of the 88 statements in  
    the pf Survey, based on the definition for this particular group type. The statement
    will be rated on a scale of -100 to 100, where -100 means least aligned with the principles of the 
    leadership type under consideration and +100 means most aligned. Additionally provide an explanation for the rating.
    
    Parameters:
    statement: The inputted Statement (one of 88 possible statements from the pf Survey).
    leader_type: The particular group type simulated to rate the statement
    group_type_description: A description of this particular group type.

    Returns:
    str: A JSON string consisting of 2 outputs... the 'rating' and 'explanation' of why that rating was chosen.
    """
    delimiter = "####"
    system_message = f"""
    Here is the definition of a {group_type}: {group_type_description}
    
    You will be provided with a statement.
    The statement will be delimited with {delimiter} characters.
    
    On a scale of -100 to +100 how would such a {group_type} rate this statement? 
    Where -100 indicates the least amount of agreement and +100 indicates the most amount of agreement.
    
    I need you to provide two pieces of output:
    1. The rating (a numerical score between -100 and +100 inclusive).
    2. An explanation of why you gave that score in 40 words or less.
    
    Please provide two outputs in the form of a valid JSON object
    {{
        "rating": <numerical_score>,
        "explanation": <reason_for_score>
    }}
    
    - The "rating" must be a number between -100 and +100 inclusive.
    - The "explanation" should be a concise justification for the given rating in 40 words or less.
    Make sure the output is a properly formatted JSON object.
    """          
    messages =  [  
    {'role':'system', 
     'content': system_message},    
    {'role':'user', 
     'content': f"{delimiter}{statement}{delimiter}"}  
    ]

    content, token_dict = get_completion_and_token_count(messages, model=MODEL, temperature=temp)
    #logging.info(token_dict)

    # Parse the JSON content
    try:
        # Convert JSON string ('content') to Python dictionary
        json_output = json.loads(content)
    
        # Access values from the JSON
        rating = json_output.get("rating")
        explanation = json_output.get("explanation")
        if explanation is None:
            explanation = "Explanation not available"
            logging.error('Explanation is None')
        else:
            explanation = explanation.strip().replace('\n', "").replace('\r', "").replace('|', "")  
    except json.JSONDecodeError as e:
        rating = 0
        explanation = "Explanation not available"
        logging.error('JSON Decode Error')

    global num_plus_100, num_minus_100  # Access the global variables
    if rating == -100:
        num_minus_100 += 1

    if rating == 100:
        num_plus_100 += 1

    return (f'{statement}|{rating}|{explanation}', explanation != "Explanation not available")
    


def identify_line_issues(file_path):
    issues = []  # To store problematic lines and their issues
    
    with open(file_path, 'r', encoding='utf-8') as file:
        for line_number, line in enumerate(file, start=1):
            line = line.strip()  # Remove leading and trailing whitespace
            
            # Check the number of delimiters
            delimiter_count = line.count('|')
            if delimiter_count != 2:
                issues.append((line_number, line, "Incorrect number of delimiters"))
                continue
            
            # Split the line
            parts = line.split('|')
            
            # Check for empty fields
            if any(part.strip() == "" for part in parts):
                issues.append((line_number, line, "Empty field detected"))
            
            # Check for unexpected characters (e.g., control characters)
            if any('\n' in part or '\r' in part for part in parts):
                issues.append((line_number, line, "Unexpected line break or control character in fields"))
            
            # Check for leading/trailing whitespace
            if any(part != part.strip() for part in parts):
                issues.append((line_number, line, "Field contains leading or trailing whitespace"))
    
    return issues


def copy_text_file(input_file, output_file):
    try:
        with open(input_file, 'r', encoding='utf-8') as infile:
            with open(output_file, 'w', encoding='utf-8') as outfile:
                for line in infile:
                    outfile.write(line)
        logging.info(f"File successfully copied from {input_file} to {output_file}.")
    except FileNotFoundError:
        logging.error(f"Error: The file {input_file} does not exist.")
    except Exception as e:
        logging.error(f"An error occurred: {e}")

def create_folder_if_not_exists(folder_path):
    """
    Creates a folder if it doesn't already exist.

    Args:
        folder_path (str): The path of the folder to create.
    """
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"Folder created: {folder_path}")
    else:
        print(f"Folder already exists: {folder_path}")


def simulate_group_type_responses(experiment_num, num_simulations):
    """   
    Simulate how 12 different group types would evaluate each of the 88 statements from the 
    pf Survey based on the provided definition for each particular Leadership Type. Each statement
    will be rated on a scale of -100 to 100, where -100 means least aligned with the principles of the particular
    leadership type under consideration and +100 means most aligned. Additionally provide a rationale for each rating.

    For each group type, run this 'num_simulations' times  (setting the temperature to 1.4)

    File 'pf_group_themes.txt' contains the list of the 12 different group types and their descriptions
    File 'pf_88_statements.txt' contains the list of the 88 Statements in the pf Survey
    """

    txt_files_folder_name = f'txt_files/E{experiment_num}'
    create_folder_if_not_exists(txt_files_folder_name)

    xlsx_files_folder_name = f'data/E{experiment_num}_pf_simulated_responses_raw'
    create_folder_if_not_exists(xlsx_files_folder_name)

    leader_type_count = 1

    with open('pf_group_themes.txt', 'r') as group_types_file:   
        for line in group_types_file:
                split_line = line.split('|')
                group_type = split_line[0]
                group_type_description = split_line[1]
            
                for i in range(1, num_simulations + 1):
                    group_type_plus = group_type + " S" + str(i)
                    print(f'{group_type_plus}   Total Simulations Count: {group_type_count} ')
                    logging.info(f'{group_type_plus}   Total Simulations Count: {group_type_count} ')
                    group_type_count += 1
                    
                    if os.path.exists('results.txt'):
                        os.remove('results.txt')
            
                    statement_num = 0
            
                    results_string = "Statement|Rating|Rationale\n"
                    
                    with open('pf_88_statements.txt', 'r') as statement_file:
                        for statement in statement_file:
                            statement_num += 1
                            if statement_num % 10 == 0:
                                logging.info(statement_num)
                            for _ in range(10):
                                result, success = get_score_and_rationale_using_gpt(statement.strip(), group_type, group_type_description, temp=TEMP14)
                                if success:
                                    break
                                else:
                                    logging.error("Result Problem")
                            results_string += result + "\n"
                                
                    with open('results.txt', 'a') as results_file:
                        results_file.write(results_string.strip())
                        
                    issues = identify_line_issues('results.txt')
                    if issues:
                        logging.error(f"Found {len(issues)} issues:")
                        for line_number, line, issue in issues:
                            logging.error(f"Line {line_number}: {issue} -> {line}")
                    else:
                        logging.info("No issues found.")

                    txt_file_name = f'{txt_files_folder_name}/{group_type_plus}.txt'
                    txt_file_name = txt_file_name.replace(' ', '_')
                    copy_text_file('results.txt', txt_file_name)
            
                    df = pd.read_csv('results.txt', delimiter='|')
                    xlsx_file_name = f'{xlsx_files_folder_name}/E{experiment_num} {group_type_plus}.xlsx'
                    xlsx_file_name = xlsx_file_name.replace(' ', '_')
                    
                    # Write the DataFrame to an Excel file
                    df.to_excel(xlsx_file_name, index=False)

                    global num_plus_100, num_minus_100  # Access the global variables
                    logging.info(f'Cumulative number of -100 Ratings:  {num_minus_100}')
                    logging.info(f'Cumulative number of +100 Ratings:  {num_plus_100}')



# For every file within a folder adjust the column width as specified for each column as well as it's horizontal Alignment
def adjust_width_and_alignment_of_all_files_in_a_folder(input_folder, output_folder):
    # Desired column widths and alignments
    column_settings = {
        'A': {'width': 86, 'alignment': 'left'},
        'B': {'width': 13, 'alignment': 'center'},
        'C': {'width': 225, 'alignment': 'left'}
    }
    
    # Ensure output folder exists
    os.makedirs(output_folder, exist_ok=True)
    
    # Iterate through all .xlsx files in the folder
    for filename in os.listdir(input_folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(input_folder, filename)
            
            # Load the workbook
            wb = load_workbook(file_path)
            
            # Adjust settings for each sheet in the workbook
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                for col, settings in column_settings.items():
                    # Set column width
                    sheet.column_dimensions[col].width = settings['width']
                    
                    # Set horizontal alignment for all cells in the column
                    for cell in sheet[col]:
                        cell.alignment = Alignment(horizontal=settings['alignment'])
            
            # Save the updated workbook
            output_file_path = os.path.join(output_folder, filename)
            wb.save(output_file_path)
            print(f"Processed {filename} and saved to {output_file_path}")
            logging.info(f"Processed {filename} and saved to {output_file_path}")



def get_args():
    parser = ap.ArgumentParser(description='Command Line parser written in python',
                                prog = 'pf_survey_simulated_responses.py',)

    parser.add_argument('--ex_num', '-e',
                       type = str,
                       help='Experiment Number',
                       required=True)
    parser.add_argument('--sims', '-s',
                    type = int,
                    help='Number of Simulations per Group Type',
                    required=True)

    return parser.parse_args()


def main():
    args = get_args()
    setup_logging(args.ex_num)
    simulate_group_type_responses(args.ex_num, args.sims)

    input_folder_name = f'data/E{args.ex_num}_pf_simulated_responses_raw'
    output_folder_name = f'data/E{args.ex_num}_PF_Simulated_Responses'
    adjust_width_and_alignment_of_all_files_in_a_folder(input_folder_name, output_folder_name)


if __name__ == "__main__":
    main()